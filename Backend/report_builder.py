# report_builder.py
# Excel report generator for the VivoDPI QA audit.
#
# One row = one issue or warning. Columns are tuned for a dev team that will
# reproduce and fix the bugs, so we keep:
#   - the annotated screenshot (so the dev sees exactly where the defect is)
#   - text (OCR), suggestion, both explanations (ES/EN)
#   - severity, category, type, both confidence values
#   - region IDs and filtered-by-pre, so the dev can rewind into the raw OCR
#   - source path, kept but hidden by default (path is bulky, name shown instead)
#
# Removed (compared to earlier versions):
#   - "Original" column (redundant with Annotated)
#   - "Screen" full path column (basename is enough; full path lives hidden)
#   - bbox X/Y/W/H columns (the rectangle in the image already shows them)
#   - "Img W/H" columns (image is embedded, dimensions are visible)
#   - "Pre-Analysis Tag" column (legacy field, not used by the dev team)

from io import BytesIO
from pathlib import Path

import cv2
import numpy as np

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .logger import get_logger


log = get_logger("REPORT")


# Palette: vivid blues/greens for a dev-team report
PALETTE = {
    "header_bg":      "1E88E5",   # vivid blue
    "header_fg":      "FFFFFF",

    # row fills by type (mild backgrounds so chips still pop)
    "row_issue":      "FDECEA",   # soft red
    "row_warning":    "FFF8E1",   # pale yellow

    # severity chips: thermometer scale (red -> orange -> yellow -> green)
    # so the gravity reads instantly. Yellow text for "low" because the
    # background is light; everything else uses white text.
    "sev_high_bg":    "D32F2F",   # vivid red (danger)
    "sev_high_fg":    "FFFFFF",
    "sev_medium_bg":  "F57C00",   # strong orange (warning)
    "sev_medium_fg":  "FFFFFF",
    "sev_low_bg":     "FBC02D",   # yellow (caution)
    "sev_low_fg":     "424242",   # dark grey: yellow + white is unreadable
    "sev_info_bg":    "388E3C",   # green (not an issue, just noise)
    "sev_info_fg":    "FFFFFF",

    # type chips
    "type_issue_bg":  "C62828",
    "type_warn_bg":   "F9A825",

    # confidence chips
    "conf_high":      "2E7D32",
    "conf_medium":    "F9A825",
    "conf_low":       "9E9E9E",

    "text_dim":       "78909C",
    "border":         "CFD8DC",
    "white":          "FFFFFF",
}


# Thumbnail target dimensions in pixels; row/col sized so image fits inside.
# THUMB_MAX_HEIGHT_PX caps very tall screenshots (scrolling captures) so a
# single row never balloons to monstrous heights. The thumbnail is always
# scaled DOWN to fit (THUMB_WIDTH_PX, THUMB_MAX_HEIGHT_PX), aspect ratio kept.
THUMB_WIDTH_PX = 540
THUMB_MAX_HEIGHT_PX = 800
# Row height in Excel points (1 pt ~= 1.33 px); some padding so the image
# doesn't touch the cell borders.
ROW_HEIGHT_PT = int(THUMB_MAX_HEIGHT_PX * 0.75) + 8
# Image column width in Excel units (~= px / 7), small padding
IMG_COL_WIDTH = int(THUMB_WIDTH_PX / 7) + 2


# Each entry: (key, header, width, alignment, kind, hidden)
COLUMNS = [
    ("type",            "Type",             12,  "center", "type",       False),
    ("severity",        "Severity",         11,  "center", "severity",   False),
    ("source_name",     "Screenshot",       26,  "left",   "text",       False),
    ("img_annotated",   "Annotated",        IMG_COL_WIDTH, "center", "img", False),
    ("category",        "Category",         18,  "left",   "category",   False),
    ("text_excerpt",    "OCR Text",         32,  "left",   "text",       False),
    ("suggestion",      "Suggestion",       32,  "left",   "text",       False),
    ("explanation_es",  "Explanation (ES)", 42,  "left",   "text",       False),
    ("explanation_en",  "Explanation (EN)", 42,  "left",   "text",       False),
    ("llm_confidence",  "LLM Confidence",   13,  "center", "confidence", False),
    ("ocr_trust",       "OCR Trust",        11,  "center", "num",        False),
    ("region_ids",      "Region IDs",       14,  "center", "list",       False),
    ("filtered_by_pre", "Pre-Filtered",     13,  "center", "bool",       False),
    ("source_path",     "Source Path",      60,  "left",   "text",       True),
]


def _read_image(path: str):
    try:
        img = cv2.imread(path)
        if img is None:
            log.warn("Could not read image: %s", path)
        return img
    except Exception as e:
        log.warn("Exception reading %s: %s", path, e)
        return None


def _annotate_bbox(img: np.ndarray, bbox: dict) -> np.ndarray:
    """Solid red rectangle (BGR 0,0,255), 4 px thick."""
    if img is None:
        return None
    annotated = img.copy()
    try:
        x = int(bbox.get("x") or 0)
        y = int(bbox.get("y") or 0)
        w = int(bbox.get("width") or 0)
        h = int(bbox.get("height") or 0)
        if w <= 0 or h <= 0:
            return annotated
        cv2.rectangle(annotated, (x, y), (x + w, y + h), (0, 0, 255), 4)
    except Exception as e:
        log.warn("Failed to draw bbox %s: %s", bbox, e)
    return annotated


def _thumbnail_bytes(img: np.ndarray):
    """Resize so it fits inside (THUMB_WIDTH_PX, THUMB_MAX_HEIGHT_PX)."""
    h, w = img.shape[:2]
    scale = THUMB_WIDTH_PX / w
    new_w = THUMB_WIDTH_PX
    new_h = int(h * scale)
    if new_h > THUMB_MAX_HEIGHT_PX:
        scale = THUMB_MAX_HEIGHT_PX / h
        new_h = THUMB_MAX_HEIGHT_PX
        new_w = int(w * scale)
    thumb = cv2.resize(img, (new_w, new_h), interpolation=cv2.INTER_AREA)
    ok, buf = cv2.imencode(".png", thumb)
    if not ok:
        raise RuntimeError("cv2.imencode failed")
    bio = BytesIO(buf.tobytes())
    bio.seek(0)
    return bio, new_w, new_h


def _xlimage_from(img: np.ndarray) -> XLImage:
    bio, w, h = _thumbnail_bytes(img)
    xlimg = XLImage(bio)
    xlimg.width = w
    xlimg.height = h
    return xlimg


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _border_all(color_hex: str = PALETTE["border"]) -> Border:
    s = Side(style="thin", color=color_hex)
    return Border(left=s, right=s, top=s, bottom=s)


_HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color=PALETTE["header_fg"])
_BODY_FONT = Font(name="Segoe UI", size=10)
_BODY_FONT_MONO = Font(name="Cascadia Code", size=9)
_DIM_FONT = Font(name="Segoe UI", size=9, color=PALETTE["text_dim"], italic=True)
_CHIP_FONT = Font(name="Segoe UI", size=10, bold=True, color=PALETTE["white"])


def build_rows(records: list) -> list:
    """Pass-through for dict records; legacy list format converted on the fly."""
    out = []
    for r in records:
        out.append(r if isinstance(r, dict) else _record_list_to_dict(r))
    return out


def _record_list_to_dict(rec: list) -> dict:
    return {
        "type":            rec[0],
        "source_name":     Path(str(rec[1])).name,
        "source_path":     str(rec[1]),
        "category":        rec[2],
        "text_excerpt":    rec[3],
        "suggestion":      rec[4],
        "explanation_es":  rec[5],
        "explanation_en":  "",
        "severity":        rec[6],
        "ocr_trust":       rec[7],
        "llm_confidence":  rec[8],
        "pre_analysis":    rec[9],
        "bbox_x":          rec[11],
        "bbox_y":          rec[12],
        "bbox_w":          rec[13],
        "bbox_h":          rec[14],
        "image_w":         rec[15],
        "image_h":         rec[16],
        "filtered_by_pre": rec[17],
    }


def export_excel(rows: list, filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "QA Report"
    ws.sheet_view.showGridLines = False

    _write_header(ws)
    _write_body(ws, rows)
    _apply_filters_and_freeze(ws, len(rows))
    _apply_conditional_formats(ws, len(rows))
    _hide_columns(ws)

    wb.save(filename)
    log.info("Report saved: %s (%d rows)", filename, len(rows))


def _write_header(ws):
    ws.row_dimensions[1].height = 32
    for idx, (_, header, width, align, kind, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _fill(PALETTE["header_bg"])
        cell.alignment = Alignment(
            horizontal=align if align != "left" else "left",
            vertical="center",
            wrap_text=True,
        )
        cell.border = _border_all(PALETTE["header_bg"])
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_body(ws, rows):
    for row_idx, record in enumerate(rows, start=2):
        ws.row_dimensions[row_idx].height = ROW_HEIGHT_PT

        row_type = (record.get("type") or "").upper()
        row_bg = (
            PALETTE["row_issue"] if row_type == "ISSUE" else
            PALETTE["row_warning"] if row_type == "WARNING" else
            None
        )

        for col_idx, (key, _, _, align, kind, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = record.get(key)

            if kind == "img":
                _place_image(ws, row_idx, col_idx, record)
                if row_bg:
                    cell.fill = _fill(row_bg)
                continue

            cell.value = _format_value(value, kind)
            cell.alignment = Alignment(
                horizontal=align, vertical="center", wrap_text=True,
            )
            cell.border = _border_all()

            if row_bg:
                cell.fill = _fill(row_bg)

            if key in ("text_excerpt", "suggestion"):
                cell.font = _BODY_FONT_MONO
            elif kind == "num" and (value is None or value == ""):
                cell.font = _DIM_FONT
            else:
                cell.font = _BODY_FONT

            if kind == "severity":
                _color_severity_chip(cell, value)
            elif kind == "type":
                _color_type_chip(cell, value)
            elif kind == "confidence":
                _color_confidence_chip(cell, value)

            if key == "ocr_trust" and isinstance(value, (int, float)):
                cell.number_format = '0.00'


def _format_value(value, kind):
    if value is None:
        return ""
    if kind == "bool":
        return "Yes" if value else "No"
    if kind == "list":
        if isinstance(value, list):
            return ", ".join(str(v) for v in value)
        return str(value)
    return value


def _color_severity_chip(cell, value):
    v = str(value or "").strip().lower()
    chip_map = {
        "high":   (PALETTE["sev_high_bg"],   PALETTE["sev_high_fg"]),
        "medium": (PALETTE["sev_medium_bg"], PALETTE["sev_medium_fg"]),
        "low":    (PALETTE["sev_low_bg"],    PALETTE["sev_low_fg"]),
        "info":   (PALETTE["sev_info_bg"],   PALETTE["sev_info_fg"]),
    }
    if v in chip_map:
        bg, fg = chip_map[v]
        cell.fill = _fill(bg)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color=fg)


def _color_type_chip(cell, value):
    v = str(value or "").strip().upper()
    if v == "ISSUE":
        cell.fill = _fill(PALETTE["type_issue_bg"])
        cell.font = _CHIP_FONT
    elif v == "WARNING":
        cell.fill = _fill(PALETTE["type_warn_bg"])
        cell.font = _CHIP_FONT


def _color_confidence_chip(cell, value):
    v = str(value or "").strip().lower()
    color_map = {
        "high":   PALETTE["conf_high"],
        "medium": PALETTE["conf_medium"],
        "low":    PALETTE["conf_low"],
    }
    if v in color_map:
        cell.fill = _fill(color_map[v])
        cell.font = _CHIP_FONT


def _place_image(ws, row_idx, col_idx, record):
    """
    Embed the annotated thumbnail INSIDE the cell using a two-cell anchor.
    Two-cell anchor pins the image to both the top-left and bottom-right
    corners of the target cell, so the image moves and resizes WITH the cell
    when the user filters, sorts, or resizes rows/columns.
    """
    src_path = record.get("source_path")
    if not src_path:
        cell = ws.cell(row=row_idx, column=col_idx, value="(no image)")
        cell.font = _DIM_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return

    img = _read_image(src_path)
    if img is None:
        cell = ws.cell(row=row_idx, column=col_idx, value="(missing)")
        cell.font = _DIM_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return

    bbox = {
        "x":      record.get("bbox_x"),
        "y":      record.get("bbox_y"),
        "width":  record.get("bbox_w"),
        "height": record.get("bbox_h"),
    }
    if any(bbox[k] is None for k in ("x", "y", "width", "height")):
        payload = img
    else:
        payload = _annotate_bbox(img, bbox)

    try:
        xlimg = _xlimage_from(payload)
        # Two-cell anchor: image is bound to the rectangle between
        # (col_idx, row_idx) and (col_idx+1, row_idx+1), i.e. inside this
        # single cell. openpyxl uses 0-based indices for AnchorMarker.
        from openpyxl.drawing.spreadsheet_drawing import (
            AnchorMarker, TwoCellAnchor,
        )
        # 0-based col/row for AnchorMarker
        col0 = col_idx - 1
        row0 = row_idx - 1
        marker_from = AnchorMarker(col=col0,     colOff=0, row=row0,     rowOff=0)
        marker_to   = AnchorMarker(col=col0 + 1, colOff=0, row=row0 + 1, rowOff=0)
        xlimg.anchor = TwoCellAnchor(
            editAs="oneCell",   # image resizes with cell, doesn't free-float
            _from=marker_from,
            to=marker_to,
        )
        ws.add_image(xlimg)
    except Exception as e:
        log.warn("Embed failed (row %d): %s", row_idx, e)


def _apply_filters_and_freeze(ws, n_rows):
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{n_rows + 1}"
    ws.freeze_panes = "A2"


def _apply_conditional_formats(ws, n_rows):
    sev_col_idx = next(
        i for i, (k, *_) in enumerate(COLUMNS, start=1) if k == "severity"
    )
    sev_letter = get_column_letter(sev_col_idx)
    rng = f"{sev_letter}2:{sev_letter}{n_rows + 1}"
    rules = [
        ("high",   PALETTE["sev_high_bg"]),
        ("medium", PALETTE["sev_medium_bg"]),
        ("low",    PALETTE["sev_low_bg"]),
        ("info",   PALETTE["sev_info_bg"]),
    ]
    for value, color in rules:
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="equal",
                formula=[f'"{value}"'],
                fill=_fill(color),
            ),
        )


def _hide_columns(ws):
    for idx, (_, _, _, _, _, hidden) in enumerate(COLUMNS, start=1):
        if hidden:
            ws.column_dimensions[get_column_letter(idx)].hidden = True